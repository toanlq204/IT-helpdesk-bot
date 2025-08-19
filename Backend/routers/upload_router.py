from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.responses import JSONResponse
from services.upload_file_service import UploadFileService
from typing import Optional
import io
import os
from pathlib import Path

# File processing imports
import PyPDF2
import docx
import openpyxl
import csv
import json as json_lib

router = APIRouter()

# Dependency injection
def get_upload_service():
    return UploadFileService()

# Supported file types and their MIME types
SUPPORTED_FILE_TYPES = {
    'text/plain': ['.txt', '.md', '.log', '.conf', '.ini'],
    'application/pdf': ['.pdf'],
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document': ['.docx'],
    'application/msword': ['.doc'],
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': ['.xlsx'],
    'application/vnd.ms-excel': ['.xls'],
    'text/csv': ['.csv'],
    'application/json': ['.json'],
    'application/xml': ['.xml'],
    'text/xml': ['.xml'],
    'text/html': ['.html', '.htm']
}

def extract_text_from_file(file_content: bytes, filename: str, content_type: str) -> str:
    """
    Extract text content from various file types
    """
    file_extension = Path(filename).suffix.lower()
    
    try:
        # Text files
        if content_type.startswith('text/') or file_extension in ['.txt', '.md', '.log', '.conf', '.ini', '.html', '.htm', '.xml']:
            return file_content.decode('utf-8', errors='ignore')
        
        # PDF files
        elif file_extension == '.pdf':
            pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            return text
        
        # Word documents
        elif file_extension == '.docx':
            doc = docx.Document(io.BytesIO(file_content))
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        
        # Excel files
        elif file_extension in ['.xlsx', '.xls']:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
                text += "\n"
            return text
        
        # CSV files
        elif file_extension == '.csv':
            csv_content = file_content.decode('utf-8', errors='ignore')
            return csv_content
        
        # JSON files
        elif file_extension == '.json':
            json_content = file_content.decode('utf-8', errors='ignore')
            try:
                # Pretty print JSON for better readability
                parsed_json = json_lib.loads(json_content)
                return json_lib.dumps(parsed_json, indent=2)
            except:
                return json_content
        
        else:
            # Try to decode as text for unknown types
            return file_content.decode('utf-8', errors='ignore')
            
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Failed to extract text from file: {str(e)}"
        )

def validate_file_type(filename: str, content_type: str) -> bool:
    """
    Validate if the file type is supported
    """
    file_extension = Path(filename).suffix.lower()
    
    # Check by content type
    for supported_type, extensions in SUPPORTED_FILE_TYPES.items():
        if content_type.startswith(supported_type) or file_extension in extensions:
            return True
    
    # Allow common text-based files
    text_extensions = ['.txt', '.md', '.log', '.conf', '.ini', '.yaml', '.yml', '.json', '.xml', '.html', '.htm']
    if file_extension in text_extensions:
        return True
    
    return False

@router.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    use_ai_metadata: bool = Form(True),
    chunk_size: int = Form(2000),
    custom_metadata: Optional[str] = Form(None),
    upload_service: UploadFileService = Depends(get_upload_service)
):
    """
    Upload a file, extract its text content, and store it with AI-generated metadata
    
    - **file**: The file to upload (supports txt, pdf, docx, xlsx, csv, json, xml, html, etc.)
    - **use_ai_metadata**: Whether to use AI for metadata generation (default: True)
    - **chunk_size**: Size of chunks for AI analysis (default: 2000)
    - **custom_metadata**: Optional custom metadata as JSON string
    """
    try:
        # Validate file size (max 50MB)
        max_file_size = 50 * 1024 * 1024  # 50MB
        if file.size and file.size > max_file_size:
            raise HTTPException(
                status_code=413,
                detail=f"File too large. Maximum size is {max_file_size // (1024*1024)}MB"
            )
        
        # Validate file type
        if not validate_file_type(file.filename, file.content_type):
            supported_extensions = []
            for extensions in SUPPORTED_FILE_TYPES.values():
                supported_extensions.extend(extensions)
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type. Supported formats: {', '.join(set(supported_extensions))}"
            )
        
        # Read file content
        file_content = await file.read()
        
        # Extract text from file
        text_content = extract_text_from_file(file_content, file.filename, file.content_type)
        
        # Validate extracted content
        if not text_content.strip():
            raise HTTPException(
                status_code=400,
                detail="No text content could be extracted from the file"
            )
        
        # Parse custom metadata if provided
        metadata = None
        if custom_metadata:
            try:
                metadata = json_lib.loads(custom_metadata)
            except json_lib.JSONDecodeError:
                raise HTTPException(
                    status_code=400,
                    detail="Invalid JSON format for custom_metadata"
                )
        
        # Store file content with metadata
        result = upload_service.store_file_content(
            file_content=text_content,
            file_name=file.filename,
            metadata=metadata,
            use_ai_metadata=use_ai_metadata
        )
        
        if result["status"] == "error":
            raise HTTPException(
                status_code=500,
                detail=f"Failed to store file: {result['message']}"
            )
        
        # Return success response with metadata
        return JSONResponse(
            status_code=200,
            content={
                "status": "success",
                "message": f"File '{file.filename}' uploaded and processed successfully",
                "file_info": {
                    "filename": file.filename,
                    "content_type": file.content_type,
                    "file_size": len(file_content),
                    "text_length": len(text_content),
                    "word_count": len(text_content.split())
                }
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Unexpected error during file upload: {str(e)}"
        )

@router.get("/upload/supported-types")
async def get_supported_file_types():
    """
    Get list of supported file types and formats
    """
    all_extensions = []
    for extensions in SUPPORTED_FILE_TYPES.values():
        all_extensions.extend(extensions)
    
    return {
        "supported_extensions": sorted(set(all_extensions)),
        "file_types": {
            "documents": [".pdf", ".docx", ".doc"],
            "spreadsheets": [".xlsx", ".xls", ".csv"],
            "text_files": [".txt", ".md", ".log", ".conf", ".ini"],
            "data_files": [".json", ".xml", ".csv"],
            "web_files": [".html", ".htm"]
        },
        "max_file_size": "50MB",
        "features": [
            "AI-powered metadata extraction",
            "Content chunking for large files",
            "Automatic text extraction",
            "Custom metadata support"
        ]
    }

@router.post("/upload/bulk")
async def bulk_upload_files(
    files: list[UploadFile] = File(...),
    use_ai_metadata: bool = Form(True),
    chunk_size: int = Form(2000),
    upload_service: UploadFileService = Depends(get_upload_service)
):
    """
    Upload multiple files at once
    
    - **files**: List of files to upload
    - **use_ai_metadata**: Whether to use AI for metadata generation
    - **chunk_size**: Size of chunks for AI analysis
    """
    if len(files) > 10:
        raise HTTPException(
            status_code=400,
            detail="Maximum 10 files allowed per bulk upload"
        )
    
    results = []
    errors = []
    
    for file in files:
        try:
            # Process each file individually
            file_content = await file.read()
            text_content = extract_text_from_file(file_content, file.filename, file.content_type)
            
            result = upload_service.store_file_content(
                file_content=text_content,
                file_name=file.filename,
                use_ai_metadata=use_ai_metadata
            )
            
            results.append({
                "filename": file.filename,
                "status": result["status"],
                "metadata": result.get("metadata", {})
            })
            
        except Exception as e:
            errors.append({
                "filename": file.filename,
                "error": str(e)
            })
    
    return {
        "total_files": len(files),
        "successful_uploads": len(results),
        "failed_uploads": len(errors),
        "results": results,
        "errors": errors
    } 